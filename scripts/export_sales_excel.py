#!/usr/bin/env python3
import json
import os
import re
import sys
import urllib.request
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage


PROJECT_ROOT = Path(__file__).resolve().parents[1]
INPUT_PATH = PROJECT_ROOT / "output" / "sales-data.json"
OUTPUT_PATH = PROJECT_ROOT / "output" / "sales-data.xlsx"
IMAGE_DIR = PROJECT_ROOT / "output" / "excel-images"
WPS_UPLOAD_IMAGE_MAX_SIZE = int(os.environ.get("WPS_UPLOAD_IMAGE_MAX_SIZE") or os.environ.get("WPS_UPLOAD_IMAGE_SIZE") or "140")

HEADERS = ["图片", "SPU", "今日销量", "流量日期", "曝光量", "点击量", "名字"]


def main():
    source_excel_path = source_excel_from_args()
    data = json.loads(INPUT_PATH.read_text(encoding="utf-8"))
    source_rows = read_source_excel(source_excel_path) if source_excel_path else None
    rows = merge_rows_by_spu(data, source_rows)

    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SPU数据"
    setup_sheet(sheet)

    for index, row in enumerate(rows, start=2):
        sheet.cell(index, 2).value = row["spuId"]
        sheet.cell(index, 3).value = as_number(row.get("todaySales"))
        sheet.cell(index, 4).value = row.get("trafficDate") or ""
        sheet.cell(index, 5).value = as_number(row.get("exposure"))
        sheet.cell(index, 6).value = as_number(row.get("clicks"))
        sheet.cell(index, 7).value = row.get("name") or ""

        for column in range(2, 8):
            sheet.cell(index, column).alignment = Alignment(horizontal="center", vertical="center")

        sheet.row_dimensions[index].height = 62
        image_path = row.get("sourceImagePath")
        if not image_path and not source_rows:
            image_path = download_and_prepare_image(row.get("imageSrc"), row["spuId"])
        if image_path:
            image = ExcelImage(str(image_path))
            image.width = 72
            image.height = 72
            sheet.add_image(image, f"A{index}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Saved Excel: {OUTPUT_PATH}")
    print(f"Rows: {len(rows)}")


def source_excel_from_args():
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).expanduser().resolve()
    env_path = os.environ.get("SOURCE_EXCEL_PATH")
    if env_path:
        return Path(env_path).expanduser().resolve()
    return None


def read_source_excel(path):
    if not path.exists():
        raise FileNotFoundError(f"Input Excel not found: {path}")

    image_by_dispimg_id = extract_cell_images(path)
    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    spu_column = source_spu_column()
    source_rows = []
    seen_spus = set()

    for row_index in range(1, sheet.max_row + 1):
        spu_id = normalize_spu(sheet.cell(row_index, spu_column).value)
        if not spu_id or spu_id in seen_spus:
            continue

        seen_spus.add(spu_id)
        name = sheet.cell(row_index, 2).value or ""
        image_id = extract_dispimg_id(sheet.cell(row_index, 3).value)
        source_rows.append({
            "spuId": spu_id,
            "name": str(name).strip(),
            "sourceImagePath": image_by_dispimg_id.get(image_id),
        })

    return source_rows


def source_spu_column():
    value = os.environ.get("SOURCE_EXCEL_SPU_COLUMN", "A")
    return column_name_to_index(value)


def column_name_to_index(value):
    text = str(value).strip().upper()
    if text.isdigit():
        column = int(text)
        if column < 1:
            raise ValueError("SOURCE_EXCEL_SPU_COLUMN must be 1 or greater.")
        return column

    column = 0
    for char in text:
        if not ("A" <= char <= "Z"):
            raise ValueError(f"Invalid SOURCE_EXCEL_SPU_COLUMN: {value}")
        column = column * 26 + (ord(char) - ord("A") + 1)
    if column < 1:
        raise ValueError("SOURCE_EXCEL_SPU_COLUMN cannot be empty.")
    return column


def extract_cell_images(path):
    image_paths = {}
    with ZipFile(path) as archive:
        names = set(archive.namelist())
        if "xl/cellimages.xml" not in names or "xl/_rels/cellimages.xml.rels" not in names:
            return image_paths

        rel_targets = parse_cell_image_relationships(archive.read("xl/_rels/cellimages.xml.rels"))
        cell_images_root = ET.fromstring(archive.read("xl/cellimages.xml"))
        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }

        for picture in cell_images_root.findall(".//xdr:pic", ns):
            name_element = picture.find(".//xdr:cNvPr", ns)
            blip_element = picture.find(".//a:blip", ns)
            image_id = name_element.attrib.get("name") if name_element is not None else ""
            rel_id = blip_element.attrib.get(f"{{{ns['r']}}}embed") if blip_element is not None else ""
            target = rel_targets.get(rel_id)
            if not image_id or not target or target == "NULL":
                continue

            archive_path = f"xl/{target}"
            if archive_path not in names:
                continue

            suffix = Path(target).suffix or ".png"
            output_path = IMAGE_DIR / f"source-{image_id}{suffix}"
            if not output_path.exists():
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_bytes(archive.read(archive_path))
            image_paths[image_id] = prepare_local_image(output_path, image_id)

    return image_paths


def parse_cell_image_relationships(raw_xml):
    root = ET.fromstring(raw_xml)
    rel_targets = {}
    for rel in root:
      rel_id = rel.attrib.get("Id")
      target = rel.attrib.get("Target")
      if rel_id and target:
          rel_targets[rel_id] = target
    return rel_targets


def extract_dispimg_id(value):
    if value is None:
        return ""
    match = re.search(r'DISPIMG\("([^"]+)"', str(value))
    return match.group(1) if match else ""


def normalize_spu(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text if re.fullmatch(r"\d+", text) else ""


def merge_rows_by_spu(data, source_rows=None):
    sales_by_spu = {}
    traffic_by_spu = {}
    order = []

    for sale in data.get("records", []):
        spu_id = str(sale.get("spuId") or "").strip()
        if not spu_id:
            continue
        if spu_id not in order:
            order.append(spu_id)
        sales_by_spu[spu_id] = sale.get("todaySales", "")

    for traffic in data.get("trafficAnalysis", {}).get("records", []):
        spu_id = str(traffic.get("spuId") or "").strip()
        if not spu_id:
            continue
        if spu_id not in order:
            order.append(spu_id)
        traffic_by_spu.setdefault(spu_id, []).append(
            {
                "trafficDate": traffic.get("date", ""),
                "exposure": traffic.get("exposure", ""),
                "clicks": traffic.get("clicks", ""),
                "imageSrc": traffic.get("imageSrc", ""),
            }
        )

    if source_rows is None:
        rows = []
        for spu_id in order:
            base = {"spuId": spu_id, "todaySales": sales_by_spu.get(spu_id, "")}
            traffic_rows = traffic_by_spu.get(spu_id) or [{}]
            for traffic_row in traffic_rows:
                rows.append({**base, **traffic_row})
        return rows

    rows = []
    for source_row in source_rows:
        spu_id = source_row["spuId"]
        traffic_rows = traffic_by_spu.get(spu_id) or [{}]
        for traffic_row in traffic_rows:
            row = {
                **source_row,
                "spuId": spu_id,
                "todaySales": sales_by_spu.get(spu_id, ""),
                **traffic_row,
            }
            row["name"] = source_row.get("name") or ""
            row["sourceImagePath"] = source_row.get("sourceImagePath")
            rows.append(row)
    return rows


def setup_sheet(sheet):
    sheet.append(HEADERS)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [13, 16, 12, 14, 12, 12, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:G1"


def download_and_prepare_image(url, spu_id):
    if not url:
        return None

    safe_spu = re.sub(r"[^0-9A-Za-z_-]+", "_", spu_id)
    output_path = IMAGE_DIR / f"{safe_spu}.png"
    if output_path.exists():
        return output_path

    try:
        request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = response.read()
        save_prepared_image(BytesIO(raw), output_path)
        return output_path
    except Exception as error:
        print(f"Image skipped for SPU {spu_id}: {error}")
        return None


def prepare_local_image(path, image_id):
    output_path = IMAGE_DIR / f"{re.sub(r'[^0-9A-Za-z_-]+', '_', image_id)}.png"
    if output_path.exists():
        return output_path

    save_prepared_image(path, output_path)
    return output_path


def save_prepared_image(source, output_path):
    with PillowImage.open(source) as image:
        image = image.convert("RGB")
        if WPS_UPLOAD_IMAGE_MAX_SIZE > 0:
            image.thumbnail((WPS_UPLOAD_IMAGE_MAX_SIZE, WPS_UPLOAD_IMAGE_MAX_SIZE))
        image.save(output_path, "PNG")


def as_number(value):
    if value is None or value == "":
        return ""
    text = str(value).replace(",", "").strip()
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)
    return value


if __name__ == "__main__":
    main()
