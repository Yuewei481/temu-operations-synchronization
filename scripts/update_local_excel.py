#!/usr/bin/env python3
import json
import os
import re
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PAYLOAD_PATH = Path(
    os.environ.get("WPS_UPDATE_PAYLOAD")
    or os.environ.get("WPS_APPEND_PAYLOAD")
    or PROJECT_ROOT / "output" / "wps-append-payload.json"
).expanduser().resolve()


def main():
    target_path = target_excel_path()
    payload = json.loads(PAYLOAD_PATH.read_text(encoding="utf-8"))
    payload_rows = unique_payload_rows(payload.get("rows") or [])
    if not payload_rows:
        raise ValueError(f"No rows to write in {PAYLOAD_PATH}")

    # Loading from memory prevents openpyxl from retaining a Windows file
    # handle that would block the atomic replacement below.
    workbook = load_workbook(BytesIO(target_path.read_bytes()))
    sheet_name = str(os.environ.get("LOCAL_TARGET_EXCEL_SHEET_NAME") or "").strip()
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f'Local target sheet not found: "{sheet_name}"')
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    date_column = column_name_to_index(os.environ.get("LOCAL_TARGET_DATE_COLUMN") or "A")
    name_column = column_name_to_index(os.environ.get("LOCAL_TARGET_NAME_COLUMN") or "B")
    sales_column = column_name_to_index(os.environ.get("LOCAL_TARGET_SALES_COLUMN") or "C")
    start_row = positive_integer(os.environ.get("LOCAL_TARGET_START_ROW") or "2", "LOCAL_TARGET_START_ROW")
    ensure_distinct_columns(date_column, name_column, sales_column)

    existing_rows, last_used_row = index_existing_rows(
        sheet,
        start_row,
        date_column,
        name_column,
        sales_column,
    )
    updated = 0
    appended = 0
    next_row = max(start_row, last_used_row + 1)

    for row in payload_rows:
        date = normalize_date(row.get("date"))
        name = normalize_name(row.get("name"))
        if not date or not name:
            continue

        key = (date, name)
        row_index = existing_rows.get(key)
        if row_index is None:
            row_index = next_writable_row(
                sheet,
                next_row,
                date_column,
                name_column,
                sales_column,
            )
            next_row = row_index + 1
            sheet.cell(row_index, date_column).value = date
            sheet.cell(row_index, name_column).value = name
            existing_rows[key] = row_index
            appended += 1
        else:
            updated += 1

        sheet.cell(row_index, sales_column).value = number_or_text(row.get("sales"))

    active_sheet_title = sheet.title
    temporary_path = target_path.with_name(f".{target_path.stem}.tmp{target_path.suffix}")
    workbook.save(temporary_path)
    workbook.close()
    temporary_path.replace(target_path)

    print(f"Updated local Excel: {target_path}")
    print(f"Sheet: {active_sheet_title}")
    print(f"Matched rows updated: {updated}")
    print(f"New rows appended: {appended}")


def target_excel_path():
    value = str(os.environ.get("LOCAL_TARGET_EXCEL_PATH") or "").strip()
    if not value:
        raise ValueError("LOCAL_TARGET_EXCEL_PATH is required.")
    path = Path(value).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Local target Excel not found: {path}")
    return path


def unique_payload_rows(rows):
    unique = {}
    for row in rows:
        date = normalize_date(row.get("date"))
        name = normalize_name(row.get("name"))
        if date and name:
            unique[(date, name)] = row
    return group_rows_by_date(list(unique.values()))


def group_rows_by_date(rows):
    groups = {}
    for row in rows:
        date = normalize_date(row.get("date"))
        groups.setdefault(date, []).append(row)
    return [row for date in sorted(groups, key=date_sort_key) for row in groups[date]]


def date_sort_key(value):
    match = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", str(value or ""))
    if not match:
        return (1, str(value or ""))
    return (0, int(match.group(1)), int(match.group(2)), int(match.group(3)))


def index_existing_rows(sheet, start_row, date_column, name_column, sales_column):
    index = {}
    last_used_row = start_row - 1
    for row_index in range(start_row, sheet.max_row + 1):
        date = normalize_date(sheet.cell(row_index, date_column).value)
        name = normalize_name(sheet.cell(row_index, name_column).value)
        sales = sheet.cell(row_index, sales_column).value
        if date or name or sales not in (None, ""):
            last_used_row = row_index
        if row_has_merged_target_cell(
            sheet,
            row_index,
            date_column,
            name_column,
            sales_column,
        ):
            continue
        if date and name and (date, name) not in index:
            index[(date, name)] = row_index
    return index, last_used_row


def next_writable_row(sheet, start_row, *columns):
    row_index = start_row
    while row_has_merged_target_cell(sheet, row_index, *columns):
        row_index += 1
    return row_index


def row_has_merged_target_cell(sheet, row_index, *columns):
    return any(
        merged_range.min_row <= row_index <= merged_range.max_row
        and any(merged_range.min_col <= column <= merged_range.max_col for column in columns)
        for merged_range in sheet.merged_cells.ranges
    )


def normalize_date(value):
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = re.sub(r"[年月.\\-]", "/", text).replace("日", "")
    normalized = re.sub(r"/+", "/", normalized)
    match = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", normalized)
    if not match:
        return normalized
    return f"{int(match.group(1))}/{int(match.group(2))}/{int(match.group(3))}"


def normalize_name(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def number_or_text(value):
    text = str(value if value is not None else "").strip().replace(",", "")
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    return int(number) if number.is_integer() else number


def column_name_to_index(value):
    text = str(value or "").strip().upper()
    if text.isdigit():
        return positive_integer(text, "column")
    if not re.fullmatch(r"[A-Z]+", text):
        raise ValueError(f"Invalid Excel column: {value}")
    result = 0
    for character in text:
        result = result * 26 + ord(character) - ord("A") + 1
    return result


def positive_integer(value, name):
    try:
        result = int(str(value).strip())
    except ValueError as error:
        raise ValueError(f"{name} must be a positive integer: {value}") from error
    if result < 1:
        raise ValueError(f"{name} must be a positive integer: {value}")
    return result


def ensure_distinct_columns(*columns):
    if len(set(columns)) != len(columns):
        raise ValueError("Local target date, name, and sales columns must be different.")


if __name__ == "__main__":
    main()
