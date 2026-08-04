#!/usr/bin/env python3
import json
import os
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
INPUT_PATH = Path(
    os.environ.get("SALES_DATA_JSON_PATH") or PROJECT_ROOT / "output" / "sales-data.json"
).expanduser().resolve()
OUTPUT_PATH = Path(
    os.environ.get("SKU_SALES_EXCEL_PATH") or PROJECT_ROOT / "output" / "sku-sales.xlsx"
).expanduser().resolve()


def main():
    data = json.loads(INPUT_PATH.read_text(encoding="utf-8"))
    name_by_sku_cargo = read_name_lookup(source_excel_path())
    sales_by_date = collect_sales_by_date(data, name_by_sku_cargo)
    if not sales_by_date:
        raise ValueError(f"No dated SKU sales records found in {INPUT_PATH}")

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sales_date, rows in sales_by_date.items():
        sheet = workbook.create_sheet(title=safe_sheet_title(sales_date))
        sheet.append(["名字", "销量"])
        style_sheet(sheet)

        for name, sales in rows.items():
            sheet.append([name, as_number(sales)])
            sheet.cell(sheet.max_row, 1).alignment = Alignment(horizontal="center")
            sheet.cell(sheet.max_row, 2).alignment = Alignment(horizontal="center")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Saved SKU sales Excel: {OUTPUT_PATH}")
    print(f"Date sheets: {len(sales_by_date)}")
    print(f"Rows by date: {', '.join(f'{date}={len(rows)}' for date, rows in sales_by_date.items())}")


def source_excel_path():
    value = str(os.environ.get("SOURCE_EXCEL_PATH") or "").strip()
    if not value:
        raise ValueError("SOURCE_EXCEL_PATH is required to match SKU货号 to product names.")
    path = Path(value).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Source Excel not found: {path}")
    return path


def read_name_lookup(path):
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook.active
    primary_column = column_name_to_index(
        os.environ.get("SOURCE_EXCEL_SKU_CARGO_PRIMARY_COLUMN") or "A",
        "SOURCE_EXCEL_SKU_CARGO_PRIMARY_COLUMN",
    )
    secondary_column = column_name_to_index(
        os.environ.get("SOURCE_EXCEL_SKU_CARGO_SECONDARY_COLUMN") or "B",
        "SOURCE_EXCEL_SKU_CARGO_SECONDARY_COLUMN",
    )
    name_column = column_name_to_index(
        os.environ.get("SOURCE_EXCEL_NAME_COLUMN") or "C",
        "SOURCE_EXCEL_NAME_COLUMN",
    )
    lookup = {}

    # Scan the complete primary column first so it always has priority over
    # a matching number in the secondary column.
    for cargo_column in (primary_column, secondary_column):
        for row_index in range(1, sheet.max_row + 1):
            sku_cargo_no = normalize_sku_cargo_no(sheet.cell(row_index, cargo_column).value)
            name = str(sheet.cell(row_index, name_column).value or "").strip()
            if sku_cargo_no and name and sku_cargo_no not in lookup:
                lookup[sku_cargo_no] = name

    workbook.close()
    return lookup


def column_name_to_index(value, env_name):
    text = str(value or "").strip().upper()
    if text.isdigit():
        index = int(text)
        if index > 0:
            return index
    else:
        index = 0
        for char in text:
            if not "A" <= char <= "Z":
                raise ValueError(f"Invalid {env_name}: {value}")
            index = index * 26 + ord(char) - ord("A") + 1
        if index > 0:
            return index
    raise ValueError(f"Invalid {env_name}: {value}")


def collect_sales_by_date(data, name_by_sku_cargo):
    sales_by_date = OrderedDict()

    for record in data.get("records") or []:
        cargo_values = record.get("skuCargoNos") or []
        if not cargo_values and record.get("skuCargoNo"):
            cargo_values = [record.get("skuCargoNo")]
        if not cargo_values:
            # Compatibility with JSON collected before SKU货号 became the
            # canonical identifier. New collection results always use skuCargoNos.
            cargo_values = record.get("skcCargoNos") or []
        sku_cargo_numbers = unique_normalized_cargo_numbers(cargo_values)
        if not sku_cargo_numbers:
            continue

        dated_sales = record.get("salesByDate") or []
        if not dated_sales and record.get("todaySales") not in (None, ""):
            fallback_date = data.get("salesDate") or data.get("targetDate")
            if fallback_date:
                dated_sales = [{"date": fallback_date, "sales": record.get("todaySales")}]

        for sales_row in dated_sales:
            sales_date = str(sales_row.get("date") or "").strip()
            if not sales_date:
                continue
            date_rows = sales_by_date.setdefault(sales_date, OrderedDict())
            for sku_cargo_no in sku_cargo_numbers:
                name = name_by_sku_cargo.get(sku_cargo_no)
                if name and name not in date_rows:
                    date_rows[name] = sales_row.get("sales", 0)

    return sales_by_date


def unique_normalized_cargo_numbers(values):
    result = []
    seen = set()
    for value in values or []:
        normalized = normalize_sku_cargo_no(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    return result


def normalize_sku_cargo_no(value):
    if value is None:
        return ""
    return re.sub(r"\D+", "", str(value))


def as_number(value):
    text = str(value if value is not None else "0").strip()
    try:
        number = float(text)
    except ValueError:
        return 0
    return int(number) if number.is_integer() else number


def safe_sheet_title(value):
    title = re.sub(r"[\\/*?:\[\]]", "-", str(value)).strip() or "销量"
    return title[:31]


def style_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:B1"
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 12
    sheet.row_dimensions[1].height = 24


if __name__ == "__main__":
    main()
