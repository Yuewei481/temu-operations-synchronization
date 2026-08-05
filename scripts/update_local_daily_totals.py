#!/usr/bin/env python3
import json
import os
import re
from io import BytesIO
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PAYLOAD_PATH = Path(
    os.environ.get("LOCAL_DAILY_TOTAL_PAYLOAD")
    or PROJECT_ROOT / "output" / "local-daily-sales-totals.json"
).expanduser().resolve()


def main():
    target_path = local_total_excel_path()
    payload = json.loads(PAYLOAD_PATH.read_text(encoding="utf-8"))
    rows = unique_total_rows(payload.get("rows") or [])
    if not rows:
        raise ValueError(f"No daily total rows to write in {PAYLOAD_PATH}")

    workbook = load_workbook(BytesIO(target_path.read_bytes()))
    sheet_name = required_env("LOCAL_TOTAL_SHEET_NAME")
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'Local total sheet not found: "{sheet_name}"')
    sheet = workbook[sheet_name]

    date_column = column_name_to_index(required_env("LOCAL_TOTAL_DATE_COLUMN"))
    sales_column = column_name_to_index(required_env("LOCAL_TOTAL_SALES_COLUMN"))
    start_row = positive_integer(required_env("LOCAL_TOTAL_START_ROW"), "LOCAL_TOTAL_START_ROW")
    if date_column == sales_column:
        raise ValueError("Local total date and sales columns must be different.")

    existing_rows, last_used_row = index_existing_dates(
        sheet,
        start_row,
        date_column,
        sales_column,
    )
    updated = 0
    appended = 0
    next_row = max(start_row, last_used_row + 1)

    for row in rows:
        sales_date = normalize_date(row.get("date"))
        row_index = existing_rows.get(sales_date)
        if row_index is None:
            row_index = next_row
            next_row += 1
            sheet.cell(row_index, date_column).value = sales_date
            existing_rows[sales_date] = row_index
            appended += 1
        else:
            updated += 1
        sheet.cell(row_index, sales_column).value = number_or_text(row.get("sales"))

    temporary_path = target_path.with_name(f".{target_path.stem}.tmp{target_path.suffix}")
    workbook.save(temporary_path)
    workbook.close()
    temporary_path.replace(target_path)

    print(f"Updated local daily totals: {target_path}")
    print(f"Sheet: {sheet_name}")
    print(f"Existing dates updated: {updated}")
    print(f"New dates appended: {appended}")


def local_total_excel_path():
    value = str(
        os.environ.get("LOCAL_TOTAL_EXCEL_PATH")
        or os.environ.get("LOCAL_TARGET_EXCEL_PATH")
        or ""
    ).strip()
    if not value:
        raise ValueError("LOCAL_TOTAL_EXCEL_PATH is required.")
    path = Path(value).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Local total Excel not found: {path}")
    return path


def required_env(name):
    value = str(os.environ.get(name) or "").strip()
    if not value:
        raise ValueError(f"{name} is required.")
    return value


def unique_total_rows(rows):
    unique = {}
    for row in rows:
        sales_date = normalize_date(row.get("date"))
        if not sales_date:
            raise ValueError(f"Invalid local daily-total date: {row.get('date')}")
        if sales_date in unique:
            raise ValueError(f"Duplicate local daily-total date: {sales_date}")
        unique[sales_date] = row
    return [unique[key] for key in sorted(unique, key=date_sort_key)]


def index_existing_dates(sheet, start_row, date_column, sales_column):
    index = {}
    last_used_row = start_row - 1
    for row_index in range(start_row, sheet.max_row + 1):
        sales_date = normalize_date(sheet.cell(row_index, date_column).value)
        sales = sheet.cell(row_index, sales_column).value
        if sales_date or sales not in (None, ""):
            last_used_row = row_index
        if sales_date and sales_date not in index:
            index[sales_date] = row_index
    return index, last_used_row


def normalize_date(value):
    text = str(value or "").strip()
    match = re.fullmatch(r"(\d{4})[/.\-](\d{1,2})[/.\-](\d{1,2})", text)
    if not match:
        return ""
    year, month, day = (int(part) for part in match.groups())
    try:
        date(year, month, day)
    except ValueError:
        return ""
    return f"{year}/{month}/{day}"


def date_sort_key(value):
    year, month, day = (int(part) for part in value.split("/"))
    return year, month, day


def number_or_text(value):
    text = str(value if value is not None else "").strip().replace(",", "")
    if not text:
        raise ValueError("Local daily-total sales value is missing.")
    try:
        number = float(text)
    except ValueError as error:
        raise ValueError(f"Invalid local daily-total sales value: {value}") from error
    return int(number) if number.is_integer() else number


def column_name_to_index(value):
    text = str(value or "").strip().upper()
    if text.isdigit():
        return positive_integer(text, "column")
    if not re.fullmatch(r"[A-Z]+", text):
        raise ValueError(f"Invalid Excel column: {value}")
    result = 0
    for character in text:
        result = result * 26 + ord(character) - ord("A") + 1
    return result


def positive_integer(value, name):
    try:
        result = int(str(value).strip())
    except ValueError as error:
        raise ValueError(f"{name} must be a positive integer: {value}") from error
    if result < 1:
        raise ValueError(f"{name} must be a positive integer: {value}")
    return result


if __name__ == "__main__":
    main()
