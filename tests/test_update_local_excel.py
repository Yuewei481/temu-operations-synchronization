import json
import os
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

import update_local_excel


class UpdateLocalExcelTest(unittest.TestCase):
    def test_updates_existing_row_and_appends_missing_row(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            workbook_path = temporary_path / "target.xlsx"
            payload_path = temporary_path / "payload.json"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "运营数据记录表"
            sheet.append(["日期", "商品名称", "销量", "不要修改"])
            sheet.append(["2026/8/1", "商品A", 1, "保留"])
            workbook.save(workbook_path)
            workbook.close()

            payload_path.write_text(
                json.dumps(
                    {
                        "rows": [
                            {"date": "2026/8/1", "name": "商品A", "sales": 5},
                            {"date": "2026/8/2", "name": "商品A", "sales": 8},
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            environment = {
                "LOCAL_TARGET_EXCEL_PATH": str(workbook_path),
                "LOCAL_TARGET_EXCEL_SHEET_NAME": "运营数据记录表",
                "LOCAL_TARGET_DATE_COLUMN": "A",
                "LOCAL_TARGET_NAME_COLUMN": "B",
                "LOCAL_TARGET_SALES_COLUMN": "C",
                "LOCAL_TARGET_START_ROW": "2",
            }
            with patch.dict(os.environ, environment, clear=False):
                with patch.object(update_local_excel, "PAYLOAD_PATH", payload_path):
                    update_local_excel.main()

            result = load_workbook(BytesIO(workbook_path.read_bytes()), data_only=True)
            result_sheet = result["运营数据记录表"]
            self.assertEqual(result_sheet["C2"].value, 5)
            self.assertEqual(result_sheet["D2"].value, "保留")
            self.assertEqual(result_sheet["A3"].value, "2026/8/2")
            self.assertEqual(result_sheet["B3"].value, "商品A")
            self.assertEqual(result_sheet["C3"].value, 8)
            result.close()

    def test_appends_rows_in_date_batches(self):
        rows = [
            {"date": "2026/8/2", "name": "商品A", "sales": 1},
            {"date": "2026/8/1", "name": "商品B", "sales": 2},
            {"date": "2026-08-02", "name": "商品C", "sales": 3},
            {"date": "2026年8月1日", "name": "商品D", "sales": 4},
        ]

        result = update_local_excel.unique_payload_rows(rows)

        self.assertEqual([row["name"] for row in result], ["商品B", "商品D", "商品A", "商品C"])


if __name__ == "__main__":
    unittest.main()
