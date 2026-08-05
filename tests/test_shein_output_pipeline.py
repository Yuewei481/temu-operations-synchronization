import json
import os
import subprocess
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]


class SheinOutputPipelineTest(unittest.TestCase):
    def test_shared_lookup_drives_payload_and_standalone_excel(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            source_path = temporary_path / "shared-lookup.xlsx"
            sales_path = temporary_path / "shein-sales.json"
            payload_path = temporary_path / "wps-payload.json"
            daily_totals_path = temporary_path / "daily-totals.json"
            export_path = temporary_path / "shein-sales.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["PRIMARY-202501", "", "Primary Name"])
            sheet.append(["", "ALT-202502", "Secondary Name"])
            # The complete primary column must win over an earlier/later
            # secondary-column occurrence of the same normalized number.
            sheet.append(["", "SECONDARY-202501", "Wrong Secondary Name"])
            workbook.save(source_path)
            workbook.close()

            sales_path.write_text(
                json.dumps(
                    {
                        "source": "SHEIN",
                        "records": [
                            {
                                "skuCargoNo": "PP-202501",
                                "skuCargoNos": ["PP-202501"],
                                "salesByDate": [
                                    {"date": "2026-08-01", "sales": 3},
                                    {"date": "2026-08-02", "sales": 4},
                                ],
                            },
                            {
                                "skuCargoNo": "ZZ-202502",
                                "skuCargoNos": ["ZZ-202502"],
                                "salesByDate": [
                                    {"date": "2026-08-01", "sales": 5},
                                    {"date": "2026-08-02", "sales": 6},
                                ],
                            },
                            {
                                "skuCargoNo": "UNMATCHED-999999",
                                "skuCargoNos": ["UNMATCHED-999999"],
                                "salesByDate": [{"date": "2026-08-01", "sales": 99}],
                            },
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            environment = {
                **os.environ,
                "SALES_DATA_JSON_PATH": str(sales_path),
                "SOURCE_EXCEL_PATH": str(source_path),
                "SOURCE_EXCEL_SKU_CARGO_PRIMARY_COLUMN": "A",
                "SOURCE_EXCEL_SKU_CARGO_SECONDARY_COLUMN": "B",
                "SOURCE_EXCEL_NAME_COLUMN": "C",
                "SOURCE_EXCEL_IMAGE_COLUMN": "",
                "WPS_UPDATE_PAYLOAD": str(payload_path),
                "WPS_DAILY_TOTAL_PAYLOAD": str(daily_totals_path),
                "SKU_SALES_EXCEL_PATH": str(export_path),
            }

            subprocess.run(
                [sys.executable, "scripts/build_wps_append_payload.py", str(source_path)],
                cwd=PROJECT_ROOT,
                env=environment,
                check=True,
                capture_output=True,
                text=True,
            )
            payload = json.loads(payload_path.read_text(encoding="utf-8"))
            writable_rows = [
                (row["date"], row["name"], row["sales"])
                for row in payload["rows"]
            ]
            self.assertEqual(
                writable_rows,
                [
                    ("2026/8/1", "Primary Name", "3"),
                    ("2026/8/2", "Primary Name", "4"),
                    ("2026/8/1", "Secondary Name", "5"),
                    ("2026/8/2", "Secondary Name", "6"),
                ],
            )
            self.assertNotIn("Wrong Secondary Name", {row["name"] for row in payload["rows"]})

            subprocess.run(
                [sys.executable, "scripts/build_daily_sales_totals.py"],
                cwd=PROJECT_ROOT,
                env=environment,
                check=True,
                capture_output=True,
                text=True,
            )
            daily_totals = json.loads(daily_totals_path.read_text(encoding="utf-8"))
            self.assertEqual(
                daily_totals["rows"],
                [
                    {"date": "2026/8/1", "sales": 8},
                    {"date": "2026/8/2", "sales": 10},
                ],
            )
            self.assertEqual(daily_totals["diagnostics"]["products"], 2)

            subprocess.run(
                [sys.executable, "scripts/export_sku_sales_excel.py"],
                cwd=PROJECT_ROOT,
                env=environment,
                check=True,
                capture_output=True,
                text=True,
            )
            exported = load_workbook(BytesIO(export_path.read_bytes()), data_only=True)
            self.assertEqual(set(exported.sheetnames), {"2026-08-01", "2026-08-02"})
            self.assertEqual(
                list(exported["2026-08-01"].values),
                [("名字", "销量"), ("Primary Name", 3), ("Secondary Name", 5)],
            )
            self.assertEqual(
                list(exported["2026-08-02"].values),
                [("名字", "销量"), ("Primary Name", 4), ("Secondary Name", 6)],
            )
            exported.close()


if __name__ == "__main__":
    unittest.main()
