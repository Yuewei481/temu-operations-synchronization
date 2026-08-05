import json
import os
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

import update_local_daily_totals


class UpdateLocalDailyTotalsTests(unittest.TestCase):
    def test_updates_existing_date_and_appends_new_dates_in_order(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            target_path = directory / "target.xlsx"
            payload_path = directory / "daily-totals.json"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "总销量表"
            sheet["A3"] = "2026/8/1"
            sheet["B3"] = 1
            workbook.save(target_path)

            payload_path.write_text(
                json.dumps({
                    "rows": [
                        {"date": "2026/8/3", "sales": 9},
                        {"date": "2026/8/1", "sales": 7},
                        {"date": "2026/8/2", "sales": 8},
                    ]
                }),
                encoding="utf-8",
            )

            env = {
                "LOCAL_TOTAL_EXCEL_PATH": str(target_path),
                "LOCAL_TOTAL_SHEET_NAME": "总销量表",
                "LOCAL_TOTAL_DATE_COLUMN": "A",
                "LOCAL_TOTAL_SALES_COLUMN": "B",
                "LOCAL_TOTAL_START_ROW": "3",
            }
            with patch.dict(os.environ, env, clear=False):
                with patch.object(update_local_daily_totals, "PAYLOAD_PATH", payload_path):
                    update_local_daily_totals.main()

            workbook = load_workbook(BytesIO(target_path.read_bytes()), data_only=False)
            sheet = workbook["总销量表"]
            self.assertEqual(sheet["A3"].value, "2026/8/1")
            self.assertEqual(sheet["B3"].value, 7)
            self.assertEqual(sheet["A4"].value, "2026/8/2")
            self.assertEqual(sheet["B4"].value, 8)
            self.assertEqual(sheet["A5"].value, "2026/8/3")
            self.assertEqual(sheet["B5"].value, 9)
            workbook.close()

    def test_rejects_duplicate_dates(self):
        with self.assertRaisesRegex(ValueError, "Duplicate local daily-total date"):
            update_local_daily_totals.unique_total_rows([
                {"date": "2026/8/1", "sales": 1},
                {"date": "2026-08-01", "sales": 2},
            ])


if __name__ == "__main__":
    unittest.main()
